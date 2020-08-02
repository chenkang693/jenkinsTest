import openpyxl


def read_excel(filename, sheet_name):
    """读取excel数据
    :param filename: 文件名称
    :param sheet_name: sheet名称
    :return: 数据列表
    """
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    list_data = []
    for x in range(2, max_row + 1):
        dict_data = dict(
            case_id=sheet.cell(row=x, column=1).value,
            url=sheet.cell(row=x, column=5).value,
            data=sheet.cell(row=x, column=6).value,
            expected=sheet.cell(row=x, column=7).value
        )
        list_data.append(dict_data)
    return list_data


import requests


def request_func(list_data, method):
    head_not_token = {
        "X-Lemonban-Media-Type": "lemonban.v2",
        "Content-Type": "application/json"
    }
    dict_res = {}
    if method == "POST" or method == "post":
        for i in list_data:
            case_id = i["case_id"]
            url = i["url"]
            data = i["data"]
            data = eval(data)
            expected = i["expected"]
            expected = eval(expected)
            result = requests.post(url=url, json=data, headers=head_not_token)
            result_msg = result.json()["msg"]
            expected_msg = expected["msg"]
            print("预期结果为: " + expected_msg)
            print("实际结果为: " + result_msg)
            print("==================")
            if result_msg == expected_msg:
                res = "pass"
            else:
                res = "unpass"
            dict_res[case_id] = res
    return dict_res


def write_excel(filename, sheet_name, dict_res):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    for x in range(2, max_row + 1):
        case_id = sheet.cell(row=x, column=1).value
        sheet.cell(row=x, column=8,value = dict_res[case_id])
    workbook.save(filename)


excel = read_excel("test_case_api.xlsx", "login")
dict_res = request_func(excel, "post")
write_excel("test_case_api.xlsx", "login",dict_res)
